from django.shortcuts import render, redirect
from django.forms import ModelForm
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from web import models
from web.forms.server import ServerModelForm
from web.utils.pagination import Pagination
from openpyxl import load_workbook


def server_list(request):
    # 构造搜索
    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["hostname__contains"] = search_data

    queryset = models.Server.objects.filter(**data_dict)

    # 分页
    page_object = Pagination(request, queryset)
    context = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
        'search_data': search_data
    }

    return render(request, 'server_list.html', context)


@csrf_exempt
def server_add(request):
    if request.method == "GET":
        form = ServerModelForm()
        return render(request, 'form.html', {'form': form})
    # 接受用户提交的数据并进行验证
    form = ServerModelForm(data=request.POST)
    # 验证通过后保存
    if form.is_valid():
        form.save()
        # /server/list/
        return redirect('/server/list/')
    else:
        return render(request, 'form.html', {'form': form})


@csrf_exempt
def server_edit(request, nid):
    server_object = models.Server.objects.filter(id=nid).first()
    if request.method == "GET":
        form = ServerModelForm(instance=server_object)
        return render(request, 'form.html', {'form': form})
    form = ServerModelForm(data=request.POST, instance=server_object)
    # 验证通过后保存
    if form.is_valid():
        form.save()
        # /server/list/
        return redirect('/server/list/')
    else:
        return render(request, 'form.html', {'form': form})


def server_delete(request, nid):
    models.Server.objects.filter(id=nid).delete()
    return JsonResponse({"status": True})

def server_multi(request):
    """ 批量上传（Excel文件）"""
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value

        # 首先判断数据是否存在
        exists = models.Server.objects.filter(hostname=text).exists()
        # 如果不存在就添加
        if not exists:
            models.Server.objects.create(hostname=text)

    return redirect('/server/list/')
